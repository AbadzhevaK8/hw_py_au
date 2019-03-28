#!python3
# updateProduce.py - Корректирует цены в электронной таблице данных об объемах продаж.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# Типы продукции и их обновленные цены.
PRICE_UPDATES = {'Lemon': 3.07,
                 'Celery': 1.19,
                 'Garlic': 1.27}

# создать цикл по строкам и обновить цены.
for rowNum in range(2, sheet.max_row): # пропустить первую строку
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
 
